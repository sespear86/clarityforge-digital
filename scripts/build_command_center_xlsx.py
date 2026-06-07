#!/usr/bin/env python3
"""Build Command_Center_c2667b93.xlsx per ~/.grok/skills/xlsx/SKILL.md."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

OUT = Path(__file__).resolve().parents[1] / "Command_Center_c2667b93.xlsx"
TODAY = "2026-06-07"
FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HDR = Font(name=FONT, bold=True, size=11)
YELLOW = PatternFill("solid", fgColor="FFFF00")
MONEY = '$#,##0;($#,##0);"-"'
PCT = "0.0%"


def inp(cell, value):
    cell.value = value
    cell.font = BLUE
    cell.fill = YELLOW


def fml(cell, formula):
    cell.value = formula
    cell.font = BLACK


def lnk(cell, formula):
    cell.value = formula
    cell.font = GREEN


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    ws["A1"] = "ClarityForge Digital — Assumptions (edit yellow cells)"
    ws["A1"].font = HDR
    ws["A2"] = (
        "Source: Command_Center_c2667b93.csv/MD, PLAN.md, batch 30, 90-day status | "
        + TODAY
    )
    for i, h in enumerate(["Assumption", "Value", "Unit", "Notes"], 1):
        ws.cell(4, i, h).font = BOLD
    data = [
        ("Monthly Ad Budget ($)", 110, "$", "Winners only; 80/20 gate"),
        ("Target AOV ($)", 18, "$", "Command_Center CSV"),
        ("Etsy Effective Take Rate", 0.22, "%", "Fees + offsite ads"),
        ("Payhip Take Rate", 0.08, "%", "Owned traffic"),
        ("View-to-Sale Conversion", 0.025, "%", "Optimized listings"),
        ("New Listings / Week", 12, "#", "Generator + image_gen"),
        ("Monthly Budget Cap ($)", 200, "$", "PLAN guardrail"),
        ("Etsy Listing Fee ($)", 0.2, "$", "Per listing"),
        ("Projection Months", 3, "#", "90-day window"),
        ("Orders / Listing / Mo (Base)", 1.5, "#", "Base scenario"),
        ("Orders / Listing / Mo (Aggr)", 2.39, "#", "Aggressive"),
        ("Orders / Listing / Mo (Moon)", 2.77, "#", "Moonshot"),
        ("Ad ROAS Gate (min)", 2.5, "x", "Spend threshold"),
        ("Winner SKU Share", 0.2, "%", "Top 20% SKUs"),
        ("Winner Revenue Share", 0.8, "%", "Pareto target"),
    ]
    for i, (lab, val, unit, note) in enumerate(data):
        row = 5 + i
        ws.cell(row, 1, lab)
        inp(ws.cell(row, 2), val)
        ws.cell(row, 3, unit)
        ws.cell(row, 4, note)
        if unit == "%" and isinstance(val, float) and val <= 1:
            ws.cell(row, 2).number_format = PCT
        elif unit == "$":
            ws.cell(row, 2).number_format = MONEY
        elif unit == "x":
            ws.cell(row, 2).number_format = "0.0x"

    sn = wb.create_sheet("Snapshot")
    sn["A1"] = "Current Snapshot — c2667b93 audit fix"
    sn["A1"].font = HDR
    sn["A3"], sn["B3"] = "Metric", "Value"
    sn["A3"].font = sn["B3"].font = BOLD
    metrics = [
        ("Report Date", TODAY),
        ("AUTON ID", "c2667b93"),
        ("Local Directory", "/home/Irikash/ClarityForge_Digital"),
        ("Clone / GitHub", "sespear86/clarityforge-digital"),
        ("90-Day Goal", "$40,000 revenue (PLAN)"),
        ("Budget Cap", "=Assumptions!B11"),
        ("Phase", "Phase 1 — Days 1-14 foundation + velocity"),
        ("Seed Etsy Listings", 3),
        ("Ideas Generated", 30),
        ("Batch Listings Ready", 30),
        ("Assets (image_gen + subagents)", 26),
        ("Subagents", "Parallel batches (image, xlsx, sync)"),
        ("Legal", "Disclosure enforced in batch JSON"),
        ("90-Day Status", "Acceleration restart; xlsx + ramp fix"),
    ]
    for i, (k, v) in enumerate(metrics, 4):
        sn.cell(i, 1, k)
        if isinstance(v, str) and v.startswith("="):
            lnk(sn.cell(i, 2), v)
        else:
            sn.cell(i, 2, v)

    pr = wb.create_sheet("Projections")
    pr["A1"] = "90-Day Scenarios (Base / Aggressive / Moonshot)"
    pr["A1"].font = HDR
    hdrs = [
        "Scenario",
        "Listings",
        "Ord/Listing/Mo",
        "Months",
        "AOV",
        "Gross Rev",
        "Take Rate",
        "Net Rev",
        "Notes",
    ]
    for i, h in enumerate(hdrs, 3):
        pr.cell(3, i, h).font = BOLD
    scen = [
        ("Base", 80, "=Assumptions!B14", "PLAN $3-8k band"),
        ("Aggressive", 150, "=Assumptions!B15", "PLAN $15-25k"),
        ("Moonshot", 250, "=Assumptions!B16", "PLAN $35k+"),
    ]
    for r, (name, lst, ordref, note) in enumerate(scen, 4):
        pr.cell(r, 1, name)
        inp(pr.cell(r, 2), lst)
        lnk(pr.cell(r, 3), ordref)
        lnk(pr.cell(r, 4), "=Assumptions!B13")
        lnk(pr.cell(r, 5), "=Assumptions!B6")
        fml(pr.cell(r, 6), f"=B{r}*C{r}*E{r}*D{r}")
        lnk(pr.cell(r, 7), "=Assumptions!B7")
        fml(pr.cell(r, 8), f"=F{r}*(1-G{r})")
        pr.cell(r, 9, note)
        pr.cell(r, 2).number_format = "0"
        for c in (5, 6, 8):
            pr.cell(r, c).number_format = MONEY
        pr.cell(r, 7).number_format = PCT

    rm = wb.create_sheet("Ramp")
    rm["A1"] = "Ramp — June 2026 (current + audit fixes)"
    rm["A1"].font = HDR
    for i, h in enumerate(
        [
            "Month",
            "New Listings",
            "New Assets",
            "Cum Listings",
            "Cum Assets",
            "Orders",
            "AOV",
            "Gross",
            "Net",
            "Ad Spend",
            "Focus",
        ],
        4,
    ):
        rm.cell(4, i, h).font = BOLD
    ramp = [
        (30, 26, 8, "M1: +30 ideas, +26 assets, xlsx, local dir, subagents"),
        (30, 20, 25, "M2: 80+ listings; Pinterest; uploads"),
        (35, 15, 30, "M3: scale; ROAS-gated ads"),
    ]
    for r, (nl, na, ord_, focus) in enumerate(ramp, 5):
        rm.cell(r, 1, ["2026-06", "2026-07", "2026-08"][r - 5])
        inp(rm.cell(r, 2), nl)
        inp(rm.cell(r, 3), na)
        if r == 5:
            fml(rm.cell(r, 4), f"=B{r}")
            fml(rm.cell(r, 5), f"=C{r}")
        else:
            fml(rm.cell(r, 4), f"=D{r-1}+B{r}")
            fml(rm.cell(r, 5), f"=E{r-1}+C{r}")
        inp(rm.cell(r, 6), ord_)
        lnk(rm.cell(r, 7), "=Assumptions!B6")
        fml(rm.cell(r, 8), f"=F{r}*G{r}")
        fml(rm.cell(r, 9), f"=H{r}*(1-Assumptions!B7)")
        lnk(rm.cell(r, 10), "=Assumptions!B5")
        rm.cell(r, 11, focus)
        for c in (2, 3, 4, 5, 6):
            rm.cell(r, c).number_format = "0"
        for c in (7, 8, 9, 10):
            rm.cell(r, c).number_format = MONEY
    rm.cell(8, 1, "Quarter").font = BOLD
    fml(rm.cell(8, 2), "=SUM(B5:B7)")
    fml(rm.cell(8, 4), "=D7")
    fml(rm.cell(8, 5), "=E7")
    fml(rm.cell(8, 6), "=SUM(F5:F7)")
    fml(rm.cell(8, 8), "=SUM(H5:H7)")
    fml(rm.cell(8, 9), "=SUM(I5:I7)")
    fml(rm.cell(8, 10), "=SUM(J5:J7)")

    p = wb.create_sheet("80-20 Analysis")
    p["A1"] = "80/20 — prioritize winners for ads"
    p["A1"].font = HDR
    for i, h in enumerate(
        ["Rank", "Listing", "Mo Rev ($)", "Cum Rev", "Cum %", "Promote?"], 4
    ):
        p.cell(4, i, h).font = BOLD
    items = [
        "ADHD Task Initiation Challenge",
        "Perimenopause NS Bundle",
        "Solopreneur Content Batching",
        "Perimenopause Mood Awareness",
        "ADHD Burnout Recovery",
        "Wall Art Calm & Focus",
        "Solopreneur Energy Calendar",
        "Perimenopause Sleep",
        "ADHD Working Memory Journal",
        "Crossover Hormone-Aware",
    ]
    revs = [120, 95, 80, 55, 45, 35, 30, 25, 20, 18]
    last = 5 + len(items) - 1
    tot_row = last + 1
    for r, (title, rev) in enumerate(zip(items, revs), 5):
        p.cell(r, 1, r - 4)
        p.cell(r, 2, title)
        inp(p.cell(r, 3), rev)
        fml(p.cell(r, 4), f"=SUM($C$5:C{r})")
        fml(p.cell(r, 5), f"=D{r}/$C${tot_row}")
        fml(p.cell(r, 6), f'=IF(E{r}<=Assumptions!B19,"YES","Hold")')
        p.cell(r, 3).number_format = MONEY
        p.cell(r, 4).number_format = MONEY
        p.cell(r, 5).number_format = PCT
    p.cell(tot_row, 2, "TOTAL").font = BOLD
    fml(p.cell(tot_row, 3), f"=SUM(C5:C{last})")

    roi = wb.create_sheet("ROI Tracker")
    roi["A1"] = "ROI Tracker — ad spend vs attributed revenue"
    roi["A1"].font = HDR
    for i, h in enumerate(
        ["Week", "Channel", "Spend", "Attrib Rev", "ROAS", "Gate", "Notes"], 4
    ):
        roi.cell(4, i, h).font = BOLD
    rows = [
        ("Jun W1", "Pinterest", 0, 0, "Pre-launch"),
        ("Jun W2", "Etsy", 0, 0, "Listing prep"),
        ("Jun W3", "Pinterest", 25, 0, "Test pins"),
        ("Jun W4", "Etsy", 35, 90, "Winners"),
        ("Jul W1", "Pinterest", 50, 140, "Scale if ROAS ok"),
    ]
    for r, (wk, ch, sp, rev, note) in enumerate(rows, 5):
        roi.cell(r, 1, wk)
        roi.cell(r, 2, ch)
        inp(roi.cell(r, 3), sp)
        inp(roi.cell(r, 4), rev)
        fml(roi.cell(r, 5), f'=IF(C{r}=0,"-",D{r}/C{r})')
        fml(roi.cell(r, 6), f'=IF(C{r}=0,"n/a",IF(E{r}>=Assumptions!B17,"YES","STOP"))')
        roi.cell(r, 7, note)
        roi.cell(r, 3).number_format = MONEY
        roi.cell(r, 4).number_format = MONEY
        roi.cell(r, 5).number_format = "0.0x"
    tr = 5 + len(rows)
    roi.cell(tr, 1, "MTD").font = BOLD
    fml(roi.cell(tr, 3), f"=SUM(C5:C{tr-1})")
    fml(roi.cell(tr, 4), f"=SUM(D5:D{tr-1})")
    fml(roi.cell(tr, 5), f'=IF(C{tr}=0,"-",D{tr}/C{tr})')
    fml(roi.cell(tr, 6), f'=IF(C{tr}<=Assumptions!B11,"Within cap","OVER")')

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    print(build())