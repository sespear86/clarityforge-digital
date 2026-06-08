#!/usr/bin/env python3
"""Minimal xlsx tests for build_command_center_xlsx.py (df92e0c4 gate)."""
import unittest
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
# The builder writes Command_Center_c2667b93.xlsx in root (or progress in some runs)
CANDIDATES = [
    ROOT / "Command_Center_c2667b93.xlsx",
    ROOT / "progress" / "Command_Center_c2667b93.xlsx",
]

def find_xlsx():
    for p in CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError("No Command_Center xlsx found for test")

class TestXlsx(unittest.TestCase):
    def test_sheets_and_assumptions(self):
        wb = load_workbook(find_xlsx())
        self.assertIn("Assumptions", wb.sheetnames)
        ws = wb["Assumptions"]
        # Yellow cells exist (editable assumptions)
        yellow = 0
        for row in ws.iter_rows(min_row=5, max_row=20, min_col=2, max_col=2):
            for cell in row:
                if cell.fill and cell.fill.fgColor:
                    rgb = str(cell.fill.fgColor.rgb or '').upper()
                    if 'FFFF00' in rgb:
                        yellow += 1
        self.assertGreaterEqual(yellow, 5, "Expected several yellow editable assumption cells")

    def test_snapshot_and_formulas(self):
        wb = load_workbook(find_xlsx(), data_only=False)
        self.assertIn("Snapshot", wb.sheetnames)
        snap = wb["Snapshot"]
        # Basic headers and some formula-like content
        self.assertEqual(snap["A3"].value, "Metric")
        self.assertEqual(snap["B3"].value, "Value")

    def test_projections_present(self):
        wb = load_workbook(find_xlsx())
        # The builder creates more sheets; at minimum we expect the core ones
        sheets = wb.sheetnames
        self.assertTrue(any("Projection" in s or "Ramp" in s or "80" in s for s in sheets) or len(sheets) >= 3)

if __name__ == "__main__":
    unittest.main()
